"""
Feedback Importer: reads completed practitioner review Excel workbooks,
parses verdicts, updates feedback gate JSON, and writes to Pinecone.

This is the upload-side of the round-trip:
  review_exporter.py generates → practitioner fills → feedback_importer.py ingests

Produces:
  1. feedback_gate{N}_completed.json — structured verdicts for the deal
  2. Pinecone updates — signal verdicts written to Signal Intelligence Layer
  3. recalibration_report.json — accuracy metrics + learning signals

Why: The system can only learn if practitioner feedback actually flows back.
This module makes ingestion a single CLI command or function call — no friction.

Usage:
    python -m tools.feedback_importer --deal "Project Jewel" --gate 1 \
        --file outputs/Project\ Jewel/review_gate1.xlsx --practitioner "Shiva"
"""

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import typer
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).parent.parent / "outputs"

app = typer.Typer()


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------

def _parse_gate1_excel(filepath: str | Path) -> dict:
    """
    Parse a completed Gate 1 review Excel workbook.

    Returns:
        Dict with signal_ratings list and metadata.
    """
    wb = load_workbook(str(filepath), data_only=True)

    if "Signals" not in wb.sheetnames:
        raise ValueError(f"Expected 'Signals' sheet in {filepath}")

    ws = wb["Signals"]
    headers = [cell.value for cell in ws[1]]

    # Map header names to column indices
    col_map = {h: i for i, h in enumerate(headers) if h}

    signal_ratings = []
    reviewed_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}

        signal_id = row_dict.get("Signal ID", "")
        if not signal_id:
            continue

        verdict = (row_dict.get("Verdict") or "").strip().upper()
        if verdict and verdict not in ("CONFIRMED", "NOISE", "UNCERTAIN"):
            verdict = ""

        corrected_rating = (row_dict.get("Corrected Rating") or "").strip().upper()
        if corrected_rating and corrected_rating not in ("RED", "YELLOW", "GREEN"):
            corrected_rating = ""

        entry = {
            "signal_id": signal_id,
            "item_type": row_dict.get("Type", "signal"),
            "pillar": row_dict.get("Pillar", ""),
            "title": row_dict.get("Title", ""),
            "original_rating": row_dict.get("Rating", ""),
            "original_confidence": row_dict.get("Confidence", ""),
            "review_urgency": row_dict.get("Review Urgency", ""),
            "verdict": verdict,
            "corrected_rating": corrected_rating or None,
            "practitioner_note": row_dict.get("Practitioner Note", "") or "",
            "additional_evidence_source": row_dict.get("Additional Evidence Source", "") or "",
            "follow_up_owner": row_dict.get("Follow-up Owner", "") or "",
        }
        signal_ratings.append(entry)

        if verdict:
            reviewed_count += 1

    wb.close()

    return {
        "signal_ratings": signal_ratings,
        "reviewed_count": reviewed_count,
        "total_count": len(signal_ratings),
    }


def _parse_gate2_excel(filepath: str | Path) -> dict:
    """
    Parse a completed Gate 2 review Excel workbook.

    Returns:
        Dict with finding_ratings, blind_spot_ratings, chase_question_statuses.
    """
    wb = load_workbook(str(filepath), data_only=True)
    result = {
        "finding_ratings": [],
        "blind_spot_ratings": [],
        "chase_question_statuses": [],
        "reviewed_findings": 0,
        "total_findings": 0,
    }

    # --- Findings sheet ---
    if "Findings" in wb.sheetnames:
        ws = wb["Findings"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}

            finding_id = row_dict.get("Finding ID", "")
            if not finding_id:
                continue

            verdict = (row_dict.get("Verdict") or "").strip().upper()
            if verdict and verdict not in ("CONFIRMED", "NOISE", "UNCERTAIN"):
                verdict = ""

            adjusted_severity = (row_dict.get("Adjusted Severity") or "").strip().upper()
            if adjusted_severity and adjusted_severity not in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
                adjusted_severity = ""

            priority = (row_dict.get("Priority") or "").strip().upper()
            effort = (row_dict.get("Remediation Effort") or "").strip().upper()

            entry = {
                "finding_id": finding_id,
                "agent": row_dict.get("Agent", ""),
                "agent_confidence": row_dict.get("Agent Confidence", ""),
                "domain": row_dict.get("Domain", ""),
                "title": row_dict.get("Title", ""),
                "original_severity": row_dict.get("Severity", ""),
                "original_confidence": row_dict.get("Confidence", ""),
                "review_urgency": row_dict.get("Review Urgency", ""),
                "verdict": verdict,
                "adjusted_severity": adjusted_severity or None,
                "practitioner_note": row_dict.get("Practitioner Note", "") or "",
                "priority": priority or None,
                "remediation_effort": effort or None,
                "additional_evidence_source": row_dict.get("Additional Evidence Source", "") or "",
                "follow_up_owner": row_dict.get("Follow-up Owner", "") or "",
            }
            result["finding_ratings"].append(entry)
            result["total_findings"] += 1
            if verdict:
                result["reviewed_findings"] += 1

    # --- Blind Spots sheet ---
    if "Blind Spots" in wb.sheetnames:
        ws = wb["Blind Spots"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}
            item_id = row_dict.get("Item ID", "")
            if not item_id:
                continue
            result["blind_spot_ratings"].append({
                "item_id": item_id,
                "item_type": row_dict.get("Type", ""),
                "agent": row_dict.get("Agent", ""),
                "domain": row_dict.get("Domain", ""),
                "has_resolution_data": (row_dict.get("Has Resolution Data") or "").strip().upper(),
                "practitioner_note": row_dict.get("Practitioner Note", "") or "",
                "follow_up_owner": row_dict.get("Follow-up Owner", "") or "",
            })

    # --- Chase Questions sheet ---
    if "Chase Questions" in wb.sheetnames:
        ws = wb["Chase Questions"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}
            q_id = row_dict.get("Question ID", "")
            if not q_id:
                continue
            status = (row_dict.get("Status") or "").strip().upper()
            result["chase_question_statuses"].append({
                "question_id": q_id,
                "pillar": row_dict.get("Pillar", ""),
                "status": status,
                "answer": row_dict.get("Answer", "") or "",
                "follow_up_owner": row_dict.get("Follow-up Owner", "") or "",
            })

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Accuracy computation
# ---------------------------------------------------------------------------

def compute_accuracy_metrics(parsed_feedback: dict, gate: int) -> dict:
    """
    Compute accuracy metrics from practitioner feedback.

    Returns a recalibration report with:
      - Overall accuracy (% CONFIRMED)
      - Noise rate (% NOISE)
      - Uncertainty rate (% UNCERTAIN)
      - Rating drift (signals where practitioner corrected the rating)
      - Per-pillar accuracy breakdown
      - Per-confidence-level accuracy (did HIGH confidence signals get confirmed more?)
      - Learning signals: specific patterns to improve
    """
    if gate == 1:
        items = parsed_feedback.get("signal_ratings", [])
        id_key = "signal_id"
        rating_key = "original_rating"
        corrected_key = "corrected_rating"
        pillar_key = "pillar"
        confidence_key = "original_confidence"
    else:
        items = parsed_feedback.get("finding_ratings", [])
        id_key = "finding_id"
        rating_key = "original_severity"
        corrected_key = "adjusted_severity"
        pillar_key = "domain"
        confidence_key = "original_confidence"

    reviewed = [i for i in items if i.get("verdict")]
    total_reviewed = len(reviewed)

    if total_reviewed == 0:
        return {
            "gate": gate,
            "total_items": len(items),
            "reviewed": 0,
            "completion_pct": 0.0,
            "accuracy": None,
            "noise_rate": None,
            "uncertainty_rate": None,
            "learning_signals": ["No feedback provided yet"],
        }

    confirmed = [i for i in reviewed if i["verdict"] == "CONFIRMED"]
    noise = [i for i in reviewed if i["verdict"] == "NOISE"]
    uncertain = [i for i in reviewed if i["verdict"] == "UNCERTAIN"]

    accuracy_pct = round(len(confirmed) / total_reviewed * 100, 1)
    noise_pct = round(len(noise) / total_reviewed * 100, 1)
    uncertain_pct = round(len(uncertain) / total_reviewed * 100, 1)

    # Rating drift: items where practitioner corrected the rating/severity
    rating_drifts = []
    for item in reviewed:
        corrected = item.get(corrected_key)
        original = item.get(rating_key, "")
        if corrected and corrected != original:
            rating_drifts.append({
                "item_id": item.get(id_key, ""),
                "original": original,
                "corrected": corrected,
                "direction": _drift_direction(original, corrected, gate),
            })

    # Per-pillar accuracy
    pillar_stats: dict[str, dict] = {}
    for item in reviewed:
        p = item.get(pillar_key, "unknown")
        if p not in pillar_stats:
            pillar_stats[p] = {"confirmed": 0, "noise": 0, "uncertain": 0, "total": 0}
        pillar_stats[p]["total"] += 1
        verdict = item["verdict"]
        if verdict in pillar_stats[p]:
            pillar_stats[p][verdict.lower()] += 1

    for p in pillar_stats:
        t = pillar_stats[p]["total"]
        pillar_stats[p]["accuracy_pct"] = round(
            pillar_stats[p]["confirmed"] / t * 100, 1
        ) if t > 0 else 0

    # Per-confidence-level accuracy
    confidence_stats: dict[str, dict] = {}
    for item in reviewed:
        c = (item.get(confidence_key) or "unknown").upper()
        if c not in confidence_stats:
            confidence_stats[c] = {"confirmed": 0, "noise": 0, "uncertain": 0, "total": 0}
        confidence_stats[c]["total"] += 1
        verdict = item["verdict"]
        if verdict.lower() in confidence_stats[c]:
            confidence_stats[c][verdict.lower()] += 1

    for c in confidence_stats:
        t = confidence_stats[c]["total"]
        confidence_stats[c]["accuracy_pct"] = round(
            confidence_stats[c]["confirmed"] / t * 100, 1
        ) if t > 0 else 0

    # Generate learning signals
    learning_signals = _generate_learning_signals(
        accuracy_pct, noise_pct, rating_drifts, pillar_stats,
        confidence_stats, noise, gate,
    )

    return {
        "gate": gate,
        "total_items": len(items),
        "reviewed": total_reviewed,
        "completion_pct": round(total_reviewed / len(items) * 100, 1) if items else 0,
        "accuracy_pct": accuracy_pct,
        "noise_rate_pct": noise_pct,
        "uncertainty_rate_pct": uncertain_pct,
        "confirmed_count": len(confirmed),
        "noise_count": len(noise),
        "uncertain_count": len(uncertain),
        "rating_drifts": rating_drifts,
        "over_rated_count": sum(1 for d in rating_drifts if d["direction"] == "over_rated"),
        "under_rated_count": sum(1 for d in rating_drifts if d["direction"] == "under_rated"),
        "pillar_accuracy": pillar_stats,
        "confidence_accuracy": confidence_stats,
        "learning_signals": learning_signals,
    }


def _drift_direction(original: str, corrected: str, gate: int) -> str:
    """Determine if the system over-rated or under-rated."""
    if gate == 1:
        severity_order = {"RED": 3, "YELLOW": 2, "GREEN": 1}
    else:
        severity_order = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}
    orig_score = severity_order.get(original.upper(), 0)
    corr_score = severity_order.get(corrected.upper(), 0)
    if orig_score > corr_score:
        return "over_rated"
    elif orig_score < corr_score:
        return "under_rated"
    return "unchanged"


def _generate_learning_signals(
    accuracy_pct: float,
    noise_pct: float,
    rating_drifts: list,
    pillar_stats: dict,
    confidence_stats: dict,
    noise_items: list,
    gate: int,
) -> list[str]:
    """
    Generate human-readable learning signals from accuracy metrics.

    These are the actionable insights that tell the system what to improve.
    """
    signals = []

    # Overall accuracy assessment
    if accuracy_pct >= 85:
        signals.append(f"Strong overall accuracy ({accuracy_pct}%) — system is well-calibrated for this deal type")
    elif accuracy_pct >= 70:
        signals.append(f"Acceptable accuracy ({accuracy_pct}%) — some calibration drift, review noise patterns")
    else:
        signals.append(f"Low accuracy ({accuracy_pct}%) — significant recalibration needed, check extraction prompts")

    # Noise analysis
    if noise_pct > 30:
        signals.append(f"High noise rate ({noise_pct}%) — system is over-extracting. Tighten signal thresholds.")
    elif noise_pct > 15:
        signals.append(f"Moderate noise rate ({noise_pct}%) — some false positives to address")

    # Rating drift
    over_rated = sum(1 for d in rating_drifts if d["direction"] == "over_rated")
    under_rated = sum(1 for d in rating_drifts if d["direction"] == "under_rated")
    if over_rated > under_rated and over_rated > 2:
        signals.append(f"System tends to over-rate severity ({over_rated} items downgraded) — calibrate toward conservative ratings")
    if under_rated > over_rated and under_rated > 2:
        signals.append(f"System tends to under-rate severity ({under_rated} items upgraded) — missing critical signals")

    # Per-pillar weak spots
    for pillar, stats in pillar_stats.items():
        if stats["total"] >= 3 and stats["accuracy_pct"] < 60:
            signals.append(f"Weak accuracy in '{pillar}' ({stats['accuracy_pct']}%) — review extraction prompts for this domain")

    # Confidence calibration: are HIGH confidence items actually more accurate?
    high_acc = confidence_stats.get("HIGH", {}).get("accuracy_pct", 0)
    medium_acc = confidence_stats.get("MEDIUM", {}).get("accuracy_pct", 0)
    if high_acc > 0 and medium_acc > 0 and medium_acc > high_acc:
        signals.append(
            f"Confidence inversion: MEDIUM-confidence items ({medium_acc}%) more accurate than "
            f"HIGH-confidence ({high_acc}%) — confidence scoring needs recalibration"
        )

    # Noise patterns: which pillars/types generate the most noise?
    noise_by_pillar: dict[str, int] = {}
    for item in noise_items:
        p = item.get("pillar", item.get("domain", "unknown"))
        noise_by_pillar[p] = noise_by_pillar.get(p, 0) + 1
    for pillar, count in sorted(noise_by_pillar.items(), key=lambda x: -x[1]):
        if count >= 2:
            signals.append(f"Noise hotspot: '{pillar}' generated {count} false positives — consider sector-specific filtering")

    return signals


# ---------------------------------------------------------------------------
# Write feedback + recalibration to disk and Pinecone
# ---------------------------------------------------------------------------

def ingest_feedback(
    filepath: str | Path,
    deal_id: str,
    gate: int,
    practitioner: str,
    company_name: str | None = None,
) -> dict:
    """
    Full feedback ingestion pipeline:
      1. Parse the completed Excel workbook
      2. Compute accuracy metrics
      3. Write feedback_gate{N}_completed.json
      4. Write recalibration_report_gate{N}.json
      5. Update Pinecone signal verdicts (Gate 1 only, graceful degradation)

    Args:
        filepath: Path to the completed review Excel workbook.
        deal_id: Deal identifier (e.g., "Project Jewel").
        gate: Gate number (1 or 2).
        practitioner: Practitioner name/ID.
        company_name: Company name for output folder. Auto-detected from deal_id if None.

    Returns:
        Dict with parsed feedback, accuracy metrics, and output paths.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Feedback file not found: {filepath}")

    # Parse
    if gate == 1:
        parsed = _parse_gate1_excel(filepath)
    elif gate == 2:
        parsed = _parse_gate2_excel(filepath)
    else:
        raise ValueError(f"Invalid gate: {gate}. Must be 1 or 2.")

    # Compute accuracy
    accuracy = compute_accuracy_metrics(parsed, gate)

    # Find the deal's output directory
    company = company_name or _find_company_dir(deal_id)
    deal_dir = OUTPUT_DIR / company
    deal_dir.mkdir(parents=True, exist_ok=True)

    # Build completed feedback structure
    feedback_completed = {
        "deal_id": deal_id,
        "gate": gate,
        "practitioner_id": practitioner,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "source_file": str(filepath),
        "accuracy_metrics": accuracy,
    }

    if gate == 1:
        feedback_completed["signal_ratings"] = parsed["signal_ratings"]
        feedback_completed["phase_accuracy_score"] = accuracy.get("accuracy_pct")
    else:
        feedback_completed["finding_ratings"] = parsed["finding_ratings"]
        feedback_completed["blind_spot_ratings"] = parsed["blind_spot_ratings"]
        feedback_completed["chase_question_statuses"] = parsed["chase_question_statuses"]

    # Write feedback
    feedback_path = deal_dir / f"feedback_gate{gate}_completed.json"
    with open(feedback_path, "w", encoding="utf-8") as f:
        json.dump(feedback_completed, f, indent=2, default=str)
    logger.info("Feedback saved to %s", feedback_path)

    # Write recalibration report
    recal_report = {
        "deal_id": deal_id,
        "gate": gate,
        "practitioner": practitioner,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        **accuracy,
    }
    recal_path = deal_dir / f"recalibration_report_gate{gate}.json"
    with open(recal_path, "w", encoding="utf-8") as f:
        json.dump(recal_report, f, indent=2, default=str)
    logger.info("Recalibration report saved to %s", recal_path)

    # Update Pinecone (Gate 1 signals only — graceful degradation)
    pinecone_updated = 0
    if gate == 1:
        try:
            from tools.signal_store import update_signal_verdict
            for rating in parsed["signal_ratings"]:
                if rating.get("verdict"):
                    update_signal_verdict(
                        deal_id=deal_id,
                        signal_id=rating["signal_id"],
                        verdict=rating["verdict"],
                        corrected_rating=rating.get("corrected_rating"),
                    )
                    pinecone_updated += 1
            logger.info("Updated %d signal verdicts in Pinecone", pinecone_updated)
        except Exception as exc:
            logger.warning("Pinecone update skipped: %s", exc)

    return {
        "feedback_path": str(feedback_path),
        "recalibration_path": str(recal_path),
        "pinecone_updated": pinecone_updated,
        "accuracy": accuracy,
        "parsed": parsed,
    }


def _find_company_dir(deal_id: str) -> str:
    """Find the company output directory for a deal_id by scanning existing outputs."""
    for p in OUTPUT_DIR.iterdir():
        if not p.is_dir() or p.name.startswith("_"):
            continue
        brief_path = p / "vdr_intelligence_brief.json"
        if brief_path.exists():
            try:
                with open(brief_path) as f:
                    brief = json.load(f)
                if brief.get("deal_id") == deal_id:
                    return p.name
            except Exception:
                continue
    return deal_id  # Fallback: use deal_id as folder name


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

@app.command()
def main(
    deal: str = typer.Option(..., help="Deal ID (e.g., 'Project Jewel')"),
    gate: int = typer.Option(..., help="Gate number (1 or 2)"),
    file: str = typer.Option(..., help="Path to the completed review Excel workbook"),
    practitioner: str = typer.Option("", help="Practitioner name/ID"),
    company: str = typer.Option("", help="Company name (auto-detected if omitted)"),
) -> None:
    """
    Import practitioner feedback from a completed review Excel workbook.

    Parses verdicts, computes accuracy metrics, saves feedback JSON,
    and updates the Signal Intelligence Layer in Pinecone.

    Example:
        python -m tools.feedback_importer --deal "Project Jewel" --gate 1 \
            --file "outputs/Project Jewel/review_gate1.xlsx" --practitioner "Shiva"
    """
    result = ingest_feedback(
        filepath=file,
        deal_id=deal,
        gate=gate,
        practitioner=practitioner or "practitioner",
        company_name=company or None,
    )

    accuracy = result["accuracy"]
    typer.echo(f"\n{'='*60}")
    typer.echo(f"  Feedback Ingested — Gate {gate}")
    typer.echo(f"{'='*60}")
    typer.echo(f"  Reviewed:    {accuracy['reviewed']} / {accuracy['total_items']}")
    typer.echo(f"  Accuracy:    {accuracy.get('accuracy_pct', 'N/A')}%")
    typer.echo(f"  Noise rate:  {accuracy.get('noise_rate_pct', 'N/A')}%")
    typer.echo(f"  Uncertain:   {accuracy.get('uncertainty_rate_pct', 'N/A')}%")
    if accuracy.get("over_rated_count", 0) or accuracy.get("under_rated_count", 0):
        typer.echo(f"  Over-rated:  {accuracy.get('over_rated_count', 0)}")
        typer.echo(f"  Under-rated: {accuracy.get('under_rated_count', 0)}")
    typer.echo(f"\n  Pinecone:    {result['pinecone_updated']} verdicts synced")
    typer.echo(f"\n  Learning signals:")
    for ls in accuracy.get("learning_signals", []):
        typer.echo(f"    • {ls}")
    typer.echo(f"\n  Files:")
    typer.echo(f"    {result['feedback_path']}")
    typer.echo(f"    {result['recalibration_path']}")


if __name__ == "__main__":
    app()
