"""
DRL Parser: Extract and normalize Crosslake OOTB Due Diligence Request List from Excel.

This module parses the standardized Excel-based DRL template into a structured JSON format
that enables field-level completeness tracking, depth scoring, and version comparison.
"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _load_schema(schema_path: str = "") -> dict[str, Any]:
    """
    Load the DRL template schema that defines tab structure and field mappings.

    Args:
        schema_path: Path to drl_template_schema.json. If empty, defaults to
                    data/drl_template_schema.json relative to project root.

    Returns:
        Dictionary containing the schema definition.

    Raises:
        FileNotFoundError: If schema file does not exist.
        json.JSONDecodeError: If schema file is invalid JSON.
    """
    if not schema_path:
        # Try to find schema relative to this file or in common locations
        possible_paths = [
            Path(__file__).parent.parent / "data" / "drl_template_schema.json",
            Path.cwd() / "data" / "drl_template_schema.json",
        ]
        for p in possible_paths:
            if p.exists():
                schema_path = str(p)
                break
        if not schema_path:
            raise FileNotFoundError(
                "Schema file not found. Please provide schema_path or ensure "
                "data/drl_template_schema.json exists."
            )

    with open(schema_path, "r") as f:
        schema = json.load(f)
    logger.info(f"Loaded schema from {schema_path}")
    return schema


def _assess_depth_score(cell_value: Any, context: dict[str, Any]) -> int:
    """
    Assess depth of a response on scale 1-10.

    Scoring rules:
    - 1-2: Single word or "Yes/No" only
    - 3-4: Brief phrase, no supporting detail
    - 5-6: Sentence-level answer with some specifics
    - 7-8: Paragraph with evidence references or data points
    - 9-10: Detailed response with document references, metrics, or linked VDR files

    Args:
        cell_value: The actual content being scored.
        context: Additional context about the field (tab type, column name, etc.).

    Returns:
        Integer depth score 1-10 (0 if empty).
    """
    if not cell_value:
        return 0

    cell_str = str(cell_value).strip()
    if not cell_str:
        return 0

    # Word count and content analysis
    words = len(cell_str.split())
    has_numbers = any(c.isdigit() for c in cell_str)
    has_parentheses = "(" in cell_str or "[" in cell_str
    has_path_indicators = "/" in cell_str or "\\" in cell_str  # VDR path reference
    newline_count = cell_str.count("\n")

    # For dataroom locations specifically
    if context.get("column_name") == "Dataroom Location" and has_path_indicators:
        return 8  # VDR path reference is substantive

    # Single word / Yes-No answers
    if words == 1 or cell_str.lower() in ("yes", "no", "n/a", "tbd"):
        return 1

    # Brief phrase (2-5 words, no structure)
    if words <= 5 and not has_numbers and not has_parentheses:
        return 3

    # Sentence level (6-20 words, some structure)
    if words <= 20 and not newline_count:
        return 5

    # Paragraph with some structure (20-100 words, has numbers or parentheses)
    if words <= 100 and (has_numbers or has_parentheses):
        return 7

    # Detailed multi-paragraph or with explicit references
    if words > 100 or newline_count > 0 or has_path_indicators:
        return 9

    return 5  # Default for moderate responses


def _parse_request_list_tab(
    ws: Any, tab_schema: dict[str, Any], tab_id: str
) -> tuple[int, int, list[dict[str, Any]]]:
    """
    Parse a 'request_list' type tab (Technology).

    Each row is a field. A field is considered "FILLED" if both
    date_responded and dataroom_location are non-null.

    Args:
        ws: openpyxl worksheet object.
        tab_schema: Schema definition for this tab.
        tab_id: Tab identifier (e.g., 'TECH').

    Returns:
        Tuple of (filled_count, total_count, fields_list).
    """
    key_columns = tab_schema["key_columns"]
    col_mapping: dict[str, int] = {}

    # Find header row and map column names to indices
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, header_name in key_columns.items():
                    if header_name.lower() in cell.value.lower():
                        col_mapping[key] = cell.column
        if len(col_mapping) == len(key_columns):
            break

    if len(col_mapping) < len(key_columns):
        logger.warning(
            f"Could not map all columns for {tab_id}. "
            f"Found {len(col_mapping)}/{len(key_columns)}"
        )

    fields = []
    filled_count = 0
    field_counter = 1

    # Parse data rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Check if row is empty
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        field_data = {
            "field_id": f"{tab_id}-{field_counter:03d}",
            "function": None,
            "request": None,
            "date_requested": None,
            "date_responded": None,
            "dataroom_location": None,
            "status": "EMPTY",
            "depth_score": 0,
            "maps_to_signals": tab_schema.get("maps_to_signals", []),
        }

        # Extract values from mapped columns
        if "function" in col_mapping:
            field_data["function"] = row[col_mapping["function"] - 1].value
        if "request" in col_mapping:
            field_data["request"] = row[col_mapping["request"] - 1].value
        if "date_requested" in col_mapping:
            field_data["date_requested"] = row[col_mapping["date_requested"] - 1].value
        if "date_responded" in col_mapping:
            field_data["date_responded"] = row[col_mapping["date_responded"] - 1].value
        if "dataroom_location" in col_mapping:
            field_data["dataroom_location"] = row[col_mapping["dataroom_location"] - 1].value

        # Determine if filled: both date_responded and dataroom_location must be non-null
        date_resp = field_data["date_responded"]
        location = field_data["dataroom_location"]

        if date_resp and location:
            field_data["status"] = "ANSWERED"
            # Assess depth based on location reference and content
            field_data["depth_score"] = _assess_depth_score(
                location, {"column_name": "Dataroom Location", "tab_id": tab_id}
            )
            filled_count += 1

        fields.append(field_data)
        field_counter += 1

    return filled_count, len(fields), fields


def _parse_inventory_table_tab(
    ws: Any, tab_schema: dict[str, Any], tab_id: str
) -> tuple[int, int, list[dict[str, Any]]]:
    """
    Parse an 'inventory_table' type tab (SoftwareDevTools, SystemsSecurityInfra).

    Each row is a field. A field is "FILLED" if it has ≥3 of its key columns populated.

    Args:
        ws: openpyxl worksheet object.
        tab_schema: Schema definition for this tab.
        tab_id: Tab identifier (e.g., 'SDT', 'SSI').

    Returns:
        Tuple of (filled_count, total_count, fields_list).
    """
    key_columns = tab_schema["key_columns"]
    col_mapping: dict[str, int] = {}

    # Map columns
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, header_name in key_columns.items():
                    if header_name.lower() in cell.value.lower():
                        col_mapping[key] = cell.column
        if len(col_mapping) == len(key_columns):
            break

    if len(col_mapping) < 3:
        logger.warning(f"Insufficient column mapping for {tab_id}. Found {len(col_mapping)}")

    fields = []
    filled_count = 0
    field_counter = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        field_data = {
            "field_id": f"{tab_id}-{field_counter:03d}",
            "status": "EMPTY",
            "depth_score": 0,
            "maps_to_signals": tab_schema.get("maps_to_signals", []),
        }

        # Extract all key column values
        populated_cols = 0
        for key, col_idx in col_mapping.items():
            value = row[col_idx - 1].value
            field_data[key] = value
            if value:
                populated_cols += 1

        # Filled if ≥3 key columns populated
        if populated_cols >= 3:
            field_data["status"] = "ANSWERED"
            # Depth based on number of populated columns and content richness
            field_data["depth_score"] = min(10, 4 + (populated_cols * 2))
            filled_count += 1

        fields.append(field_data)
        field_counter += 1

    return filled_count, len(fields), fields


def _parse_financial_table_tab(
    ws: Any, tab_schema: dict[str, Any], tab_id: str
) -> tuple[int, int, list[dict[str, Any]]]:
    """
    Parse a 'financial_table' type tab (RDSpend).

    Each row is a field. A field is "FILLED" if it has ≥2 year columns populated.

    Args:
        ws: openpyxl worksheet object.
        tab_schema: Schema definition for this tab.
        tab_id: Tab identifier (e.g., 'RDS').

    Returns:
        Tuple of (filled_count, total_count, fields_list).
    """
    key_columns = tab_schema["key_columns"]
    col_mapping: dict[str, int] = {}
    year_cols = [
        "actual_2024",
        "actual_2025",
        "budget_2026",
        "ytd_2026",
        "annualized_2026",
    ]

    # Map columns
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, header_name in key_columns.items():
                    if header_name.lower() in cell.value.lower():
                        col_mapping[key] = cell.column
        if len(col_mapping) == len(key_columns):
            break

    fields = []
    filled_count = 0
    field_counter = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        field_data = {
            "field_id": f"{tab_id}-{field_counter:03d}",
            "status": "EMPTY",
            "depth_score": 0,
            "maps_to_signals": tab_schema.get("maps_to_signals", []),
        }

        # Extract all columns
        year_populated = 0
        for key, col_idx in col_mapping.items():
            value = row[col_idx - 1].value
            field_data[key] = value
            if key in year_cols and value:
                year_populated += 1

        # Filled if ≥2 year columns populated
        if year_populated >= 2:
            field_data["status"] = "ANSWERED"
            field_data["depth_score"] = min(10, 5 + year_populated)
            filled_count += 1

        fields.append(field_data)
        field_counter += 1

    return filled_count, len(fields), fields


def _parse_roster_table_tab(
    ws: Any, tab_schema: dict[str, Any], tab_id: str
) -> tuple[int, int, list[dict[str, Any]]]:
    """
    Parse a 'roster_table' type tab (CensusInput).

    Row count matters. Table is "FILLED" if row_count > 0 AND ≥80% of rows have all
    required columns populated.

    Args:
        ws: openpyxl worksheet object.
        tab_schema: Schema definition for this tab.
        tab_id: Tab identifier (e.g., 'CEN').

    Returns:
        Tuple of (filled_count, total_count, fields_list).
    """
    key_columns = tab_schema["key_columns"]
    col_mapping: dict[str, int] = {}
    required_cols = set(key_columns.keys())

    # Map columns
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, header_name in key_columns.items():
                    if header_name.lower() in cell.value.lower():
                        col_mapping[key] = cell.column
        if len(col_mapping) == len(key_columns):
            break

    # For roster tables, we treat the entire table as one "field"
    # because row_count matters more than individual fields
    rows_data = []
    row_counter = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        row_data = {}
        populated_cols = 0
        for key, col_idx in col_mapping.items():
            value = row[col_idx - 1].value
            row_data[key] = value
            if value:
                populated_cols += 1

        # Track completeness
        row_data["completeness"] = populated_cols / len(required_cols) if required_cols else 0
        rows_data.append(row_data)
        row_counter += 1

    # Table is filled if row_count > 0 AND ≥80% of rows complete
    total_rows = len(rows_data)
    if total_rows > 0:
        complete_rows = sum(
            1 for r in rows_data if r.get("completeness", 0) >= 0.8
        )
        completion_pct = (complete_rows / total_rows * 100) if total_rows > 0 else 0

        if total_rows > 0 and completion_pct >= 80:
            filled_count = 1
            depth_score = min(10, 6 + (total_rows // 10))
        else:
            filled_count = 0
            depth_score = 0
    else:
        filled_count = 0
        depth_score = 0
        completion_pct = 0.0

    fields = [
        {
            "field_id": f"{tab_id}-001",
            "row_count": total_rows,
            "completeness_pct": completion_pct,
            "status": "ANSWERED" if filled_count > 0 else "EMPTY",
            "depth_score": depth_score,
            "maps_to_signals": tab_schema.get("maps_to_signals", []),
            "rows": rows_data,
        }
    ]

    return filled_count, 1, fields


def parse_drl_excel(filepath: str, schema_path: str = "") -> dict[str, Any]:
    """
    Parse Crosslake OOTB DRL Excel file into structured JSON.

    Detects 5 standard tabs: Technology, SoftwareDevTools, SystemsSecurityInfra,
    RDSpend, CensusInput. Extracts field-level data and computes completeness and
    depth scores per tab and overall.

    Args:
        filepath: Path to the Excel DRL file.
        schema_path: Path to drl_template_schema.json. If empty, defaults to
                    data/drl_template_schema.json.

    Returns:
        Dictionary with keys:
        - deal_id: Extracted or generated ID
        - version: Version number
        - uploaded_at: ISO timestamp
        - source_filename: Name of uploaded Excel file
        - tabs: Dict mapping tab_id to tab_state (see spec)
        - overall: Overall stats and composite score

    Raises:
        FileNotFoundError: If Excel file or schema file not found.
        Exception: If Excel file cannot be parsed.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    schema = _load_schema(schema_path)
    wb = load_workbook(filepath)

    logger.info(f"Parsing DRL Excel: {filepath}")
    logger.info(f"Available tabs: {wb.sheetnames}")

    # Map Excel sheet names to schema tab IDs
    sheet_to_tab_id = {
        "Technology": "technology",
        "SoftwareDevTools": "software_dev_tools",
        "SystemsSecurityInfra": "systems_security_infra",
        "RDSpend": "rd_spend",
        "CensusInput": "census_input",
    }

    tab_id_map = {
        "technology": "TECH",
        "software_dev_tools": "SDT",
        "systems_security_infra": "SSI",
        "rd_spend": "RDS",
        "census_input": "CEN",
    }

    drl_state: dict[str, Any] = {
        "deal_id": Path(filepath).stem.split("_")[0].upper(),
        "version": 1,
        "uploaded_at": datetime.utcnow().isoformat() + "Z",
        "source_filename": os.path.basename(filepath),
        "tabs": {},
        "overall": {
            "total_fields": 0,
            "filled_fields": 0,
            "empty_fields": 0,
            "completeness_pct": 0.0,
            "depth_score": 0.0,
            "composite_score": 0.0,
            "grade": "F",
        },
    }

    total_filled = 0
    total_fields = 0
    total_depth_sum = 0
    total_depth_count = 0

    # Parse each tab
    for sheet_name, tab_id in sheet_to_tab_id.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Tab '{sheet_name}' not found in Excel")
            drl_state["tabs"][tab_id] = {
                "total_fields": 0,
                "filled_fields": 0,
                "empty_fields": 0,
                "completeness_pct": 0.0,
                "fields": [],
            }
            continue

        ws = wb[sheet_name]
        tab_schema = schema["tabs"][sheet_name]
        tab_type = tab_schema["type"]
        numeric_id = tab_id_map[tab_id]

        logger.info(f"Parsing tab {sheet_name} (type: {tab_type})")

        # Parse based on tab type
        if tab_type == "request_list":
            filled, total, fields = _parse_request_list_tab(ws, tab_schema, numeric_id)
        elif tab_type == "inventory_table":
            filled, total, fields = _parse_inventory_table_tab(ws, tab_schema, numeric_id)
        elif tab_type == "financial_table":
            filled, total, fields = _parse_financial_table_tab(ws, tab_schema, numeric_id)
        elif tab_type == "roster_table":
            filled, total, fields = _parse_roster_table_tab(ws, tab_schema, numeric_id)
        else:
            logger.error(f"Unknown tab type: {tab_type}")
            filled, total, fields = 0, 0, []

        # Compute tab-level stats
        completeness_pct = (filled / total * 100) if total > 0 else 0.0
        depth_scores = [f["depth_score"] for f in fields if f["depth_score"] > 0]
        depth_avg = sum(depth_scores) / len(depth_scores) if depth_scores else 0.0

        drl_state["tabs"][tab_id] = {
            "total_fields": total,
            "filled_fields": filled,
            "empty_fields": total - filled,
            "completeness_pct": round(completeness_pct, 1),
            "fields": fields,
        }

        total_filled += filled
        total_fields += total
        total_depth_sum += sum(depth_scores)
        total_depth_count += len(depth_scores)

    # Compute overall stats
    overall_completeness_pct = (total_filled / total_fields * 100) if total_fields > 0 else 0.0
    overall_depth_score = (
        (total_depth_sum / total_depth_count) if total_depth_count > 0 else 0.0
    )

    # Composite score: (0.5 × completeness) + (0.5 × depth_normalized_to_100)
    depth_normalized = (overall_depth_score / 10 * 100) if overall_depth_score > 0 else 0.0
    composite_score = (0.5 * overall_completeness_pct) + (0.5 * depth_normalized)

    # Grade assignment
    if composite_score >= 85:
        grade = "A"
    elif composite_score >= 70:
        grade = "B"
    elif composite_score >= 55:
        grade = "C"
    elif composite_score >= 40:
        grade = "D"
    else:
        grade = "F"

    drl_state["overall"] = {
        "total_fields": total_fields,
        "filled_fields": total_filled,
        "empty_fields": total_fields - total_filled,
        "completeness_pct": round(overall_completeness_pct, 1),
        "depth_score": round(overall_depth_score, 1),
        "composite_score": round(composite_score, 1),
        "grade": grade,
    }

    logger.info(f"Parsing complete: {total_filled}/{total_fields} fields filled, grade {grade}")

    return drl_state
