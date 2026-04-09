"""
Review Exporter: generates practitioner review Excel workbooks for Gate 1 and Gate 2.

The Excel is a round-trip interchange format:
  1. System generates pre-populated workbook with signals/findings + empty practitioner columns
  2. Practitioner fills verdict, notes, priority, follow-up fields
  3. Completed workbook is uploaded back via feedback_importer.py

Design principles:
  - One row per reviewable item, sorted by review urgency (CRITICAL first)
  - Practitioner columns are clearly separated with yellow headers
  - Data-validation dropdowns on verdict/rating/priority columns
  - Conditional formatting: RED rows for CRITICAL urgency, amber for HIGH
  - Traceability: every row links back to signal_id/finding_id + source document
  - Instructions sheet explains the review process and column definitions

Why: Practitioners who don't have tool access need a frictionless way to review
and provide feedback. Excel is the universal interchange format in PE.
"""

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


def _format_evidence_for_excel(evidence: Any, signal_lookup: dict | None = None) -> str:
    """
    Convert structured evidence (list of typed dicts) into readable multi-line text
    suitable for an Excel cell.

    When a signal_lookup is provided, resolves signal IDs to actual VDR file
    names so practitioners see "Employee Roster.xlsx" not "SIG-003".

    Handles both legacy string evidence and the new structured format:
      [{"type": "signal", "signal_id": "SIG-001", "detail": "..."},
       {"type": "document", "source_doc": "...", "excerpt": "..."},
       {"type": "prior_agent", "agent": "Riley", "finding_id": "SEC-01", "detail": "..."},
       {"type": "missing", "expected": "...", "detail": "..."},
       {"type": "inference", "detail": "..."}]

    Args:
        evidence: Raw evidence value — str, list, dict, or None.
        signal_lookup: Optional dict mapping signal_id -> signal metadata.

    Returns:
        Formatted string with one line per evidence item, max 1500 chars.
    """
    if not evidence:
        return ""
    if isinstance(evidence, str):
        return evidence[:1500]
    if isinstance(evidence, dict):
        return json.dumps(evidence, default=str)[:1500]
    if not isinstance(evidence, list):
        return str(evidence)[:1500]

    lines = []
    for ev in evidence:
        if not isinstance(ev, dict):
            lines.append(str(ev))
            continue
        ev_type = ev.get("type", "")
        detail = ev.get("detail", "")
        if ev_type == "signal":
            sig_id = ev.get("signal_id", "")
            # Resolve signal ID to source document name
            source_file = ""
            if signal_lookup and sig_id:
                sig_meta = signal_lookup.get(sig_id, {})
                source_file = sig_meta.get("source_doc", "")
            if source_file:
                lines.append(f"[Source: {source_file}] {detail}")
            else:
                lines.append(f"[Signal {sig_id}] {detail}")
        elif ev_type == "document":
            doc = ev.get("source_doc", "Unknown")
            excerpt = ev.get("excerpt", "")
            line = f"[Source: {doc}] {detail}"
            if excerpt:
                line += f' — "{excerpt[:150]}"'
            lines.append(line)
        elif ev_type == "prior_agent":
            agent = ev.get("agent", "?")
            fid = ev.get("finding_id", "")
            lines.append(f"[Agent {agent} → {fid}] {detail}")
        elif ev_type == "missing":
            expected = ev.get("expected", "")
            lines.append(f"[Missing: {expected}] {detail}")
        elif ev_type == "inference":
            lines.append(f"[Inference] {detail}")
        else:
            lines.append(detail or str(ev))

    return "\n".join(lines)[:1500]


def _format_source_signals(source_signals: Any, signal_lookup: dict | None = None) -> str:
    """
    Format source_signals into readable source document references for Excel.

    When a signal_lookup is provided (signal_id -> {source_doc, title, ...}),
    resolves opaque IDs like "SIG-006" into the actual VDR file names
    practitioners can recognize, e.g. "Q3 Financial Model.xlsx".

    Args:
        source_signals: List of signal IDs or raw value.
        signal_lookup: Optional dict mapping signal_id -> signal metadata
                       (must contain 'source_doc' key).

    Returns:
        Formatted string with source file names, or raw IDs if no lookup.
    """
    if not source_signals:
        return ""
    if not isinstance(source_signals, list):
        return str(source_signals)

    if signal_lookup:
        docs = []
        for sig_id in source_signals:
            sig = signal_lookup.get(str(sig_id), {})
            source_doc = sig.get("source_doc", "")
            if source_doc:
                docs.append(source_doc)
            else:
                docs.append(str(sig_id))
        # Deduplicate while preserving order
        seen = set()
        unique = []
        for d in docs:
            if d not in seen:
                seen.add(d)
                unique.append(d)
        return "\n".join(unique)
    else:
        return ", ".join(str(s) for s in source_signals)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
PRACTITIONER_HEADER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
PRACTITIONER_HEADER_FONT = Font(color="000000", bold=True, size=11, name="Calibri")
CRITICAL_FILL = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
BODY_FONT = Font(size=10, name="Calibri")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Where practitioner columns start (1-indexed) for each sheet type
GATE1_PRACTITIONER_COL_START = 12  # After system columns
GATE2_PRACTITIONER_COL_START = 14


def _style_header_row(ws, num_cols: int, practitioner_start: int) -> None:
    """Apply header styling: blue for system columns, gold for practitioner columns."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        if col_idx >= practitioner_start:
            cell.fill = PRACTITIONER_HEADER_FILL
            cell.font = PRACTITIONER_HEADER_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER


def _style_data_rows(ws, num_rows: int, num_cols: int, urgency_col: int) -> None:
    """Apply row fills based on review urgency and basic formatting."""
    for row_idx in range(2, num_rows + 2):
        urgency = ws.cell(row=row_idx, column=urgency_col).value or ""
        fill = None
        if urgency == "CRITICAL":
            fill = CRITICAL_FILL
        elif urgency == "HIGH":
            fill = HIGH_FILL
        elif urgency == "MEDIUM":
            fill = MEDIUM_FILL

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    """Set column widths by 1-based index."""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_dropdown(ws, col_letter: str, options: list[str], min_row: int, max_row: int) -> None:
    """Add a dropdown data validation to a column range."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.error = "Please select from the dropdown."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


# ---------------------------------------------------------------------------
# Instructions sheet
# ---------------------------------------------------------------------------

def _add_instructions_sheet(wb: Workbook, gate: int, deal_id: str, company: str) -> None:
    """Add an Instructions sheet explaining the review process."""
    ws = wb.create_sheet("Instructions", 0)

    title_font = Font(size=16, bold=True, name="Calibri", color="1F4E79")
    section_font = Font(size=12, bold=True, name="Calibri", color="1F4E79")
    body_font = Font(size=11, name="Calibri")

    row = 1
    ws.cell(row=row, column=1, value=f"Practitioner Review — Gate {gate}").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Deal: {deal_id} | Company: {company}").font = body_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = body_font
    row += 2

    ws.cell(row=row, column=1, value="How to Use This Workbook").font = section_font
    row += 1
    instructions = [
        "1. Review items are sorted by urgency: CRITICAL items first, then HIGH, MEDIUM, LOW.",
        "2. BLUE columns are system-generated — do not modify these.",
        "3. GOLD columns are for your input — fill in verdicts, notes, and assignments.",
        "4. Use the dropdown menus for Verdict, Rating, Priority, and Effort columns.",
        "5. When done, save and upload this file back to the system for recalibration.",
        "",
        "Your feedback directly improves future scans. Every verdict teaches the system",
        "what it got right and where it drifted.",
    ]
    for line in instructions:
        ws.cell(row=row, column=1, value=line).font = body_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Verdict Definitions").font = section_font
    row += 1
    verdicts = [
        ("CONFIRMED", "Signal/finding is accurate and actionable. Optionally correct the rating if severity was wrong."),
        ("NOISE", "False positive or immaterial. The AI flagged something that doesn't matter for this deal."),
        ("UNCERTAIN", "Ambiguous — you can't tell if it's real without more data. Note what additional info is needed."),
    ]
    for verdict, desc in verdicts:
        ws.cell(row=row, column=1, value=f"  {verdict}").font = Font(size=11, bold=True, name="Calibri")
        ws.cell(row=row, column=2, value=desc).font = body_font
        row += 1

    if gate == 2:
        row += 1
        ws.cell(row=row, column=1, value="Priority Levels (Gate 2 only)").font = section_font
        row += 1
        priorities = [
            ("P1", "Must address before close — deal-breaker or value-at-risk territory"),
            ("P2", "Address in first 100 days post-acquisition"),
            ("P3", "Address within first year — important but not urgent"),
            ("P4", "Nice to have — low risk, low effort"),
        ]
        for pri, desc in priorities:
            ws.cell(row=row, column=1, value=f"  {pri}").font = Font(size=11, bold=True, name="Calibri")
            ws.cell(row=row, column=2, value=desc).font = body_font
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Remediation Effort").font = section_font
        row += 1
        efforts = [
            ("S", "Small — days, <$50K, one person"),
            ("M", "Medium — weeks, $50-200K, small team"),
            ("L", "Large — months, $200K-1M, cross-functional"),
            ("XL", "Extra-large — quarters, >$1M, org-wide initiative"),
        ]
        for eff, desc in efforts:
            ws.cell(row=row, column=1, value=f"  {eff}").font = Font(size=11, bold=True, name="Calibri")
            ws.cell(row=row, column=2, value=desc).font = body_font
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80


# ---------------------------------------------------------------------------
# Gate 1: Signal Review workbook
# ---------------------------------------------------------------------------

GATE1_SIGNAL_HEADERS = [
    # System columns (blue)
    "Signal ID", "Type", "Pillar", "Title", "Rating", "Confidence",
    "Evidence Quote", "Source Document", "Deal Implication",
    "Extraction Note", "Review Urgency", "Review Reason",
    # Practitioner columns (gold)
    "Verdict", "Corrected Rating", "Practitioner Note",
    "Additional Evidence Source", "Follow-up Owner",
]

GATE1_WIDTHS = {
    1: 12, 2: 14, 3: 22, 4: 40, 5: 10, 6: 12,
    7: 50, 8: 40, 9: 40,
    10: 30, 11: 14, 12: 50,
    13: 14, 14: 16, 15: 40,
    16: 30, 17: 20,
}


def export_gate1_workbook(
    manifest: dict,
    output_path: Path | str | None = None,
) -> Path:
    """
    Generate the Gate 1 practitioner review Excel workbook.

    Args:
        manifest: Gate 1 review manifest from practitioner_review.generate_gate1_manifest().
        output_path: Where to save. Defaults to outputs/<company>/review_gate1.xlsx.

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()
    deal_id = manifest.get("deal_id", "")
    company = manifest.get("company_name", "unknown")

    # Instructions sheet
    _add_instructions_sheet(wb, gate=1, deal_id=deal_id, company=company)

    # Signals sheet
    ws = wb.create_sheet("Signals")
    for col_idx, header in enumerate(GATE1_SIGNAL_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    items = manifest.get("review_items", [])
    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("item_id", ""))
        ws.cell(row=row_idx, column=2, value=item.get("item_type", ""))
        ws.cell(row=row_idx, column=3, value=item.get("pillar", ""))
        ws.cell(row=row_idx, column=4, value=item.get("title", ""))
        ws.cell(row=row_idx, column=5, value=item.get("rating", ""))
        ws.cell(row=row_idx, column=6, value=item.get("confidence", ""))
        ws.cell(row=row_idx, column=7, value=item.get("evidence_quote", ""))
        ws.cell(row=row_idx, column=8, value=item.get("source_doc", ""))
        ws.cell(row=row_idx, column=9, value=item.get("deal_implication", ""))
        ws.cell(row=row_idx, column=10, value=item.get("extraction_note", ""))
        ws.cell(row=row_idx, column=11, value=item.get("review_urgency", ""))
        ws.cell(row=row_idx, column=12, value=item.get("review_reason", ""))
        # Practitioner columns left blank

    num_rows = len(items)
    _style_header_row(ws, len(GATE1_SIGNAL_HEADERS), GATE1_PRACTITIONER_COL_START)
    _style_data_rows(ws, num_rows, len(GATE1_SIGNAL_HEADERS), urgency_col=11)
    _set_column_widths(ws, GATE1_WIDTHS)

    # Dropdowns
    if num_rows > 0:
        max_row = num_rows + 1
        _add_dropdown(ws, "M", ["CONFIRMED", "NOISE", "UNCERTAIN"], 2, max_row)  # Verdict
        _add_dropdown(ws, "N", ["RED", "YELLOW", "GREEN"], 2, max_row)  # Corrected Rating

    # Summary sheet
    _add_summary_sheet(wb, manifest, gate=1)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save
    if output_path is None:
        output_dir = Path(__file__).parent.parent / "outputs" / company
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "review_gate1.xlsx"
    else:
        output_path = Path(output_path)

    wb.save(str(output_path))
    logger.info("Gate 1 review workbook saved to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Gate 2: Findings + Blind Spots + Chase Questions workbook
# ---------------------------------------------------------------------------

GATE2_FINDING_HEADERS = [
    # System columns (blue)
    "Finding ID", "Agent", "Agent Confidence", "Domain", "Title",
    "Severity", "Confidence", "Confidence Reason", "Evidence",
    "Source Documents", "Deal Implication", "Review Urgency", "Review Reason",
    # Practitioner columns (gold) — starts at col 14
    "Verdict", "Adjusted Severity", "Practitioner Note",
    "Priority", "Remediation Effort", "Additional Evidence Source",
    "Follow-up Owner",
]

GATE2_BLIND_SPOT_HEADERS = [
    "Item ID", "Type", "Agent", "Domain", "Description",
    "Review Urgency",
    # Practitioner columns
    "Has Resolution Data", "Practitioner Note", "Follow-up Owner",
]

GATE2_CHASE_HEADERS = [
    "Question ID", "Pillar", "Pillar Label", "Question", "Priority",
    # Practitioner columns
    "Status", "Answer", "Follow-up Owner",
]

GATE2_FINDING_WIDTHS = {
    1: 16, 2: 12, 3: 16, 4: 28, 5: 50,      # ID, Agent, Conf, Domain, Title
    6: 12, 7: 12, 8: 40, 9: 55,              # Severity, Conf, Reason, Evidence
    10: 25, 11: 40, 12: 14, 13: 50,          # Source Signals, Deal Imp, Urgency, Reason
    14: 14, 15: 16, 16: 40,                  # Verdict, Adj Severity, Note
    17: 10, 18: 18, 19: 30,                  # Priority, Remediation, Add Evidence
    20: 20,                                   # Follow-up Owner
}


def export_gate2_workbook(
    manifest: dict,
    output_path: Path | str | None = None,
    signal_lookup: dict | None = None,
) -> Path:
    """
    Generate the Gate 2 practitioner review Excel workbook.

    Args:
        manifest: Gate 2 review manifest from practitioner_review.generate_gate2_manifest().
        output_path: Where to save. Defaults to outputs/<company>/review_gate2.xlsx.
        signal_lookup: Optional dict mapping signal_id -> signal metadata from VDR brief.
                       Used to resolve opaque signal IDs to actual source document names.

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()
    deal_id = manifest.get("deal_id", "")
    company = manifest.get("company_name", "unknown")

    _add_instructions_sheet(wb, gate=2, deal_id=deal_id, company=company)

    # --- Findings sheet ---
    ws_findings = wb.create_sheet("Findings")
    for col_idx, header in enumerate(GATE2_FINDING_HEADERS, 1):
        ws_findings.cell(row=1, column=col_idx, value=header)

    findings = manifest.get("finding_items", [])
    for row_idx, item in enumerate(findings, 2):
        ws_findings.cell(row=row_idx, column=1, value=item.get("item_id", ""))
        ws_findings.cell(row=row_idx, column=2, value=item.get("agent", ""))
        ws_findings.cell(row=row_idx, column=3, value=item.get("agent_confidence", ""))
        ws_findings.cell(row=row_idx, column=4, value=item.get("domain", ""))
        ws_findings.cell(row=row_idx, column=5, value=item.get("title", ""))
        ws_findings.cell(row=row_idx, column=6, value=item.get("severity", ""))
        ws_findings.cell(row=row_idx, column=7, value=item.get("confidence", ""))
        ws_findings.cell(row=row_idx, column=8, value=item.get("confidence_reason", ""))
        # Evidence — format structured arrays into readable multi-line text
        evidence_text = _format_evidence_for_excel(item.get("evidence", ""), signal_lookup=signal_lookup)
        ws_findings.cell(row=row_idx, column=9, value=evidence_text)
        ws_findings.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        # Source Documents — resolve signal IDs to actual file names
        ws_findings.cell(row=row_idx, column=10, value=_format_source_signals(item.get("source_signals", []), signal_lookup=signal_lookup))
        ws_findings.cell(row=row_idx, column=10).alignment = Alignment(wrap_text=True, vertical="top")
        ws_findings.cell(row=row_idx, column=11, value=item.get("deal_implication", ""))
        ws_findings.cell(row=row_idx, column=12, value=item.get("review_urgency", ""))
        ws_findings.cell(row=row_idx, column=13, value=item.get("review_reason", ""))

    num_findings = len(findings)
    _style_header_row(ws_findings, len(GATE2_FINDING_HEADERS), 14)
    _style_data_rows(ws_findings, num_findings, len(GATE2_FINDING_HEADERS), urgency_col=12)
    _set_column_widths(ws_findings, GATE2_FINDING_WIDTHS)

    if num_findings > 0:
        max_row = num_findings + 1
        _add_dropdown(ws_findings, "N", ["CONFIRMED", "NOISE", "UNCERTAIN"], 2, max_row)
        _add_dropdown(ws_findings, "O", ["CRITICAL", "HIGH", "MEDIUM", "LOW"], 2, max_row)
        _add_dropdown(ws_findings, "Q", ["P1", "P2", "P3", "P4"], 2, max_row)
        _add_dropdown(ws_findings, "R", ["S", "M", "L", "XL"], 2, max_row)

    # --- Blind Spots sheet ---
    ws_blind = wb.create_sheet("Blind Spots")
    for col_idx, header in enumerate(GATE2_BLIND_SPOT_HEADERS, 1):
        ws_blind.cell(row=1, column=col_idx, value=header)

    blind_spots = manifest.get("blind_spots", [])
    for row_idx, item in enumerate(blind_spots, 2):
        ws_blind.cell(row=row_idx, column=1, value=item.get("item_id", ""))
        ws_blind.cell(row=row_idx, column=2, value=item.get("item_type", ""))
        ws_blind.cell(row=row_idx, column=3, value=item.get("agent", ""))
        ws_blind.cell(row=row_idx, column=4, value=item.get("domain", ""))
        ws_blind.cell(row=row_idx, column=5, value=item.get("description", ""))
        ws_blind.cell(row=row_idx, column=6, value=item.get("review_urgency", ""))

    num_blind = len(blind_spots)
    _style_header_row(ws_blind, len(GATE2_BLIND_SPOT_HEADERS), 7)
    _style_data_rows(ws_blind, num_blind, len(GATE2_BLIND_SPOT_HEADERS), urgency_col=6)
    _set_column_widths(ws_blind, {1: 20, 2: 18, 3: 12, 4: 25, 5: 60, 6: 14, 7: 18, 8: 40, 9: 20})

    if num_blind > 0:
        _add_dropdown(ws_blind, "G", ["YES", "NO"], 2, num_blind + 1)

    # --- Chase Questions sheet ---
    ws_chase = wb.create_sheet("Chase Questions")
    for col_idx, header in enumerate(GATE2_CHASE_HEADERS, 1):
        ws_chase.cell(row=1, column=col_idx, value=header)

    questions = manifest.get("chase_questions", [])
    for row_idx, item in enumerate(questions, 2):
        ws_chase.cell(row=row_idx, column=1, value=item.get("item_id", ""))
        ws_chase.cell(row=row_idx, column=2, value=item.get("pillar", ""))
        ws_chase.cell(row=row_idx, column=3, value=item.get("pillar_label", ""))
        ws_chase.cell(row=row_idx, column=4, value=item.get("question", ""))
        ws_chase.cell(row=row_idx, column=5, value=item.get("priority", ""))

    num_chase = len(questions)
    _style_header_row(ws_chase, len(GATE2_CHASE_HEADERS), 6)
    _style_data_rows(ws_chase, num_chase, len(GATE2_CHASE_HEADERS), urgency_col=5)
    _set_column_widths(ws_chase, {1: 16, 2: 22, 3: 30, 4: 60, 5: 10, 6: 18, 7: 60, 8: 20})

    if num_chase > 0:
        _add_dropdown(ws_chase, "F", ["SENT", "ANSWERED", "NOT_APPLICABLE"], 2, num_chase + 1)

    # Summary sheet
    _add_summary_sheet(wb, manifest, gate=2)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if output_path is None:
        output_dir = Path(__file__).parent.parent / "outputs" / company
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "review_gate2.xlsx"
    else:
        output_path = Path(output_path)

    wb.save(str(output_path))
    logger.info("Gate 2 review workbook saved to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Summary sheet (shared by both gates)
# ---------------------------------------------------------------------------

def _add_summary_sheet(wb: Workbook, manifest: dict, gate: int) -> None:
    """Add a Summary sheet with review stats and urgency breakdown."""
    ws = wb.create_sheet("Summary")
    title_font = Font(size=14, bold=True, name="Calibri", color="1F4E79")
    label_font = Font(size=11, bold=True, name="Calibri")
    value_font = Font(size=11, name="Calibri")

    summary = manifest.get("summary", {})

    row = 1
    ws.cell(row=row, column=1, value=f"Gate {gate} Review Summary").font = title_font
    row += 2

    stats = [
        ("Deal ID", manifest.get("deal_id", "")),
        ("Company", manifest.get("company_name", "")),
        ("Gate", f"{gate} — {manifest.get('gate_label', '')}"),
        ("Generated", manifest.get("generated_at", "")),
    ]

    if gate == 1:
        stats.extend([
            ("Total Signals", summary.get("signals_total", 0)),
            ("Items to Review", summary.get("total_items", 0)),
        ])
    else:
        stats.extend([
            ("Total Findings", summary.get("total_findings", 0)),
            ("Blind Spots", summary.get("total_blind_spots", 0)),
            ("Chase Questions", summary.get("total_chase_questions", 0)),
        ])

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=str(value)).font = value_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Urgency Breakdown").font = label_font
    row += 1
    for urgency, count in summary.get("urgency_distribution", {}).items():
        ws.cell(row=row, column=1, value=urgency).font = value_font
        ws.cell(row=row, column=2, value=count).font = value_font
        row += 1

    if gate == 2 and "agent_confidence_map" in summary:
        row += 1
        ws.cell(row=row, column=1, value="Agent Confidence Levels").font = label_font
        row += 1
        for agent, conf in summary.get("agent_confidence_map", {}).items():
            ws.cell(row=row, column=1, value=agent).font = value_font
            ws.cell(row=row, column=2, value=conf).font = value_font
            row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
