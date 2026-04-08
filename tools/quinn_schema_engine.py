"""
Quinn Schema Engine — Fingerprint and diff DRL templates and signal catalogs.

This module detects structural changes to the Deal Response Library (DRL) Excel template
and the signal catalog JSON, computing cryptographic hashes of schema elements and
generating migration packets that describe what changed and its impact on existing deals.

Usage:
    from tools.quinn_schema_engine import fingerprint_drl_template, fingerprint_signal_catalog, compare_fingerprints

    fp1 = fingerprint_drl_template("data/drl_template_v1.xlsx")
    fp2 = fingerprint_drl_template("data/drl_template_v2.xlsx")
    migration = compare_fingerprints(fp1, fp2)

    if migration["changes_detected"]:
        print(f"Breaking changes: {[c for c in migration['changes'] if c['impact'] == 'BREAKING']}")
"""
from __future__ import annotations

import json
import hashlib
import logging
from pathlib import Path
from typing import Optional
from datetime import datetime, timezone

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logger = logging.getLogger(__name__)


# ── Schema Fingerprinting ─────────────────────────────────────────────────────

def fingerprint_drl_template(filepath: str) -> dict:
    """
    Parse a DRL Excel template and extract its structural fingerprint.

    Reads the Excel file, extracts tab names, column headers, field counts per tab,
    and expected row counts. Computes a SHA-256 hash of the normalized structure.

    Args:
        filepath: Path to the DRL template Excel file (e.g., data/drl_template.xlsx)

    Returns:
        {
            "source": "drl_template",
            "filepath": str,
            "timestamp": ISO-8601 timestamp when fingerprinted,
            "version": "auto-detected or inferred",
            "tabs": [
                {
                    "tab_name": str,
                    "columns": [str, ...],
                    "field_count": int,
                    "expected_row_count": int
                },
                ...
            ],
            "schema_hash": SHA-256 hex digest,
            "template_stats": {
                "total_tabs": int,
                "total_columns": int,
                "total_fields": int
            }
        }

    Raises:
        FileNotFoundError: If filepath does not exist.
        ValueError: If file cannot be parsed as Excel.
    """
    if load_workbook is None:
        raise ImportError("openpyxl required for DRL template parsing. Install with: pip install openpyxl")

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"DRL template not found: {filepath}")

    try:
        workbook = load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to parse Excel file {filepath}: {e}")

    tabs = []
    total_columns = 0
    total_fields = 0

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]

        # Extract column headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())

        # Count non-empty rows (data rows)
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_count += 1

        field_count = len(headers)
        tabs.append({
            "tab_name": sheet,
            "columns": headers,
            "field_count": field_count,
            "expected_row_count": row_count
        })

        total_columns += field_count
        total_fields += field_count

    workbook.close()

    # Compute schema hash: SHA-256 of sorted tab+column structure
    schema_string = json.dumps(
        {
            "tabs": tabs,
            "column_order": [c for t in tabs for c in t["columns"]]
        },
        sort_keys=True,
        ensure_ascii=True
    )
    schema_hash = hashlib.sha256(schema_string.encode("utf-8")).hexdigest()

    return {
        "source": "drl_template",
        "filepath": str(filepath.absolute()),
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": _infer_template_version(filepath),
        "tabs": tabs,
        "schema_hash": schema_hash,
        "template_stats": {
            "total_tabs": len(tabs),
            "total_columns": total_columns,
            "total_fields": total_fields
        }
    }


def fingerprint_signal_catalog(catalog_path: str = "") -> dict:
    """
    Parse a signal catalog JSON and extract its structural fingerprint.

    Reads the signal catalog, extracts pillar IDs, signal IDs per pillar, and counts.
    Computes a SHA-256 hash of the signal structure.

    Args:
        catalog_path: Path to signal_catalog.json. If empty, uses data/signal_catalog.json.

    Returns:
        {
            "source": "signal_catalog",
            "filepath": str,
            "timestamp": ISO-8601 timestamp,
            "version": str (from catalog["version"]),
            "pillars": [
                {
                    "pillar_id": str,
                    "pillar_label": str,
                    "signal_ids": [str, ...],
                    "signal_count": int
                },
                ...
            ],
            "schema_hash": SHA-256 hex digest,
            "catalog_stats": {
                "total_pillars": int,
                "total_signals": int
            }
        }

    Raises:
        FileNotFoundError: If catalog file not found.
        ValueError: If JSON is malformed.
    """
    if not catalog_path:
        catalog_path = Path(__file__).parent.parent / "data" / "signal_catalog.json"

    catalog_path = Path(catalog_path)
    if not catalog_path.exists():
        raise FileNotFoundError(f"Signal catalog not found: {catalog_path}")

    try:
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise ValueError(f"Malformed JSON in {catalog_path}: {e}")

    pillars = []

    # Build pillar summary from signals
    pillar_map: dict[str, dict] = {}
    for signal in catalog.get("signals", []):
        pillar_id = signal.get("pillar_id")
        if pillar_id not in pillar_map:
            pillar_map[pillar_id] = {
                "pillar_id": pillar_id,
                "pillar_label": signal.get("pillar_name", ""),
                "signal_ids": []
            }
        pillar_map[pillar_id]["signal_ids"].append(signal.get("signal_id"))

    # Sort by pillar_number if available
    for signal in catalog.get("signals", []):
        pillar_id = signal.get("pillar_id")
        if pillar_id in pillar_map and "pillar_number" not in pillar_map[pillar_id]:
            pillar_map[pillar_id]["pillar_number"] = signal.get("pillar_number")

    # Convert to sorted list
    for pillar_id in sorted(pillar_map.keys()):
        p = pillar_map[pillar_id]
        pillars.append({
            "pillar_id": p["pillar_id"],
            "pillar_label": p["pillar_label"],
            "signal_ids": sorted(p["signal_ids"]),
            "signal_count": len(p["signal_ids"])
        })

    # Compute schema hash: SHA-256 of pillar + signal structure
    schema_string = json.dumps(
        {
            "pillars": pillars,
            "all_signal_ids": sorted([s.get("signal_id") for s in catalog.get("signals", [])])
        },
        sort_keys=True,
        ensure_ascii=True
    )
    schema_hash = hashlib.sha256(schema_string.encode("utf-8")).hexdigest()

    return {
        "source": "signal_catalog",
        "filepath": str(catalog_path.absolute()),
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": catalog.get("version", "unknown"),
        "pillars": pillars,
        "schema_hash": schema_hash,
        "catalog_stats": {
            "total_pillars": len(pillars),
            "total_signals": len(catalog.get("signals", []))
        }
    }


# ── Fingerprint Comparison ────────────────────────────────────────────────────

def compare_fingerprints(fp1: dict, fp2: dict) -> dict:
    """
    Compare two fingerprints (same source type) and detect structural changes.

    Analyzes differences between DRL template fingerprints or signal catalog fingerprints,
    classifying each change as BREAKING, COMPATIBLE, or DEPRECATION.

    Args:
        fp1: First fingerprint (baseline, usually "old" version)
        fp2: Second fingerprint (current, usually "new" version)

    Returns:
        {
            "from_version": str,
            "to_version": str,
            "source_type": "drl_template" | "signal_catalog",
            "timestamp": ISO-8601,
            "changes_detected": bool,
            "schema_hash_changed": bool,
            "changes": [
                {
                    "type": "FIELD_ADDED" | "FIELD_REMOVED" | "FIELD_RENAMED" | "TAB_ADDED" | "TAB_REMOVED" | "SIGNAL_ADDED" | "SIGNAL_REMOVED" | "PILLAR_REORDERED",
                    "source": "drl_template" | "signal_catalog",
                    "tab_or_pillar": str (if applicable),
                    "field_or_signal_name": str,
                    "field_or_signal_id": str (if applicable),
                    "impact": "BREAKING" | "COMPATIBLE" | "DEPRECATION",
                    "reason": str,
                    "mitigation": str (optional)
                },
                ...
            ],
            "affected_deals": [],  # populated by version registry
            "reprocessing_required": bool,
            "breaking_changes_count": int,
            "compatible_changes_count": int
        }

    Raises:
        ValueError: If fingerprints are of different source types.
    """
    if fp1.get("source") != fp2.get("source"):
        raise ValueError(
            f"Cannot compare fingerprints of different source types: "
            f"{fp1.get('source')} vs {fp2.get('source')}"
        )

    source_type = fp1.get("source")
    changes = []

    if source_type == "drl_template":
        changes = _compare_drl_fingerprints(fp1, fp2)
    elif source_type == "signal_catalog":
        changes = _compare_catalog_fingerprints(fp1, fp2)

    schema_hash_changed = fp1.get("schema_hash") != fp2.get("schema_hash")
    breaking_count = sum(1 for c in changes if c["impact"] == "BREAKING")
    compatible_count = sum(1 for c in changes if c["impact"] == "COMPATIBLE")

    return {
        "from_version": fp1.get("version", "unknown"),
        "to_version": fp2.get("version", "unknown"),
        "source_type": source_type,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "changes_detected": schema_hash_changed,
        "schema_hash_changed": schema_hash_changed,
        "changes": changes,
        "affected_deals": [],
        "reprocessing_required": breaking_count > 0,
        "breaking_changes_count": breaking_count,
        "compatible_changes_count": compatible_count
    }


def _compare_drl_fingerprints(fp1: dict, fp2: dict) -> list[dict]:
    """
    Detect changes between two DRL template fingerprints.

    Compares tab names, column counts, and field structures.
    """
    changes = []

    tabs1 = {t["tab_name"]: t for t in fp1.get("tabs", [])}
    tabs2 = {t["tab_name"]: t for t in fp2.get("tabs", [])}

    # Detect removed or renamed tabs
    for tab_name in tabs1:
        if tab_name not in tabs2:
            changes.append({
                "type": "TAB_REMOVED",
                "source": "drl_template",
                "tab_or_pillar": tab_name,
                "field_or_signal_name": "",
                "field_or_signal_id": "",
                "impact": "BREAKING",
                "reason": f"Tab '{tab_name}' was removed from the template",
                "mitigation": "Update all parsers that reference this tab. Check for data loss in existing DRL submissions."
            })

    # Detect added tabs
    for tab_name in tabs2:
        if tab_name not in tabs1:
            changes.append({
                "type": "TAB_ADDED",
                "source": "drl_template",
                "tab_or_pillar": tab_name,
                "field_or_signal_name": "",
                "field_or_signal_id": "",
                "impact": "COMPATIBLE",
                "reason": f"New tab '{tab_name}' was added to the template",
                "mitigation": "Existing deals are unaffected. New submissions will have this tab available."
            })

    # Detect field changes within tabs
    for tab_name in tabs1:
        if tab_name in tabs2:
            cols1 = set(tabs1[tab_name]["columns"])
            cols2 = set(tabs2[tab_name]["columns"])

            removed_cols = cols1 - cols2
            for col in removed_cols:
                changes.append({
                    "type": "FIELD_REMOVED",
                    "source": "drl_template",
                    "tab_or_pillar": tab_name,
                    "field_or_signal_name": col,
                    "field_or_signal_id": "",
                    "impact": "BREAKING",
                    "reason": f"Field '{col}' in tab '{tab_name}' was removed",
                    "mitigation": "Update DRL parser to skip this field. Existing data in this column will be lost on next save."
                })

            added_cols = cols2 - cols1
            for col in added_cols:
                changes.append({
                    "type": "FIELD_ADDED",
                    "source": "drl_template",
                    "tab_or_pillar": tab_name,
                    "field_or_signal_name": col,
                    "field_or_signal_id": "",
                    "impact": "COMPATIBLE",
                    "reason": f"Field '{col}' in tab '{tab_name}' was added",
                    "mitigation": "Existing DRL submissions can be reprocessed; new field will be empty."
                })

    return changes


def _compare_catalog_fingerprints(fp1: dict, fp2: dict) -> list[dict]:
    """
    Detect changes between two signal catalog fingerprints.

    Compares pillar structures and signal inventories.
    """
    changes = []

    # Extract signal IDs by pillar
    pillars1 = {p["pillar_id"]: set(p["signal_ids"]) for p in fp1.get("pillars", [])}
    pillars2 = {p["pillar_id"]: set(p["signal_ids"]) for p in fp2.get("pillars", [])}

    # Detect removed signals
    for pillar_id in pillars1:
        if pillar_id in pillars2:
            removed_signals = pillars1[pillar_id] - pillars2[pillar_id]
            for signal_id in removed_signals:
                changes.append({
                    "type": "SIGNAL_REMOVED",
                    "source": "signal_catalog",
                    "tab_or_pillar": pillar_id,
                    "field_or_signal_name": signal_id,
                    "field_or_signal_id": signal_id,
                    "impact": "BREAKING",
                    "reason": f"Signal '{signal_id}' was removed from pillar '{pillar_id}'",
                    "mitigation": "Deals may have extracted this signal. Update extraction prompts and verify signal registry."
                })

    # Detect added signals
    for pillar_id in pillars2:
        if pillar_id in pillars1:
            added_signals = pillars2[pillar_id] - pillars1[pillar_id]
            for signal_id in added_signals:
                changes.append({
                    "type": "SIGNAL_ADDED",
                    "source": "signal_catalog",
                    "tab_or_pillar": pillar_id,
                    "field_or_signal_name": signal_id,
                    "field_or_signal_id": signal_id,
                    "impact": "COMPATIBLE",
                    "reason": f"Signal '{signal_id}' was added to pillar '{pillar_id}'",
                    "mitigation": "Re-runs of affected deals can extract the new signal. Existing scans are unaffected."
                })

    # Detect new pillars
    for pillar_id in pillars2:
        if pillar_id not in pillars1:
            changes.append({
                "type": "PILLAR_REORDERED",
                "source": "signal_catalog",
                "tab_or_pillar": pillar_id,
                "field_or_signal_name": "",
                "field_or_signal_id": "",
                "impact": "COMPATIBLE",
                "reason": f"New pillar '{pillar_id}' was added to the catalog",
                "mitigation": "Existing deals are unaffected. New scans can extract from this pillar."
            })

    return changes


# ── Utility Helpers ───────────────────────────────────────────────────────────

def _infer_template_version(filepath: Path) -> str:
    """
    Attempt to infer DRL template version from filename or content.

    Returns a version string like "1.0", "2.0", or "unknown".
    """
    filename = filepath.name.lower()

    # Try to extract version from filename patterns
    if "v1" in filename or "1.0" in filename:
        return "1.0"
    elif "v2" in filename or "2.0" in filename:
        return "2.0"
    elif "v3" in filename or "3.0" in filename:
        return "3.0"

    return "unknown"


# ── Fingerprint Storage ───────────────────────────────────────────────────────

def save_fingerprints(fingerprints: dict, output_path: str = "") -> str:
    """
    Save a collection of fingerprints to JSON for later comparison.

    Args:
        fingerprints: Dict of {"drl_template": fp, "signal_catalog": fp}
        output_path: Where to save (default: outputs/_quinn_fingerprints.json)

    Returns:
        Path where fingerprints were saved.
    """
    if not output_path:
        output_path = Path(__file__).parent.parent / "outputs" / "_quinn_fingerprints.json"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fingerprints["saved_at"] = datetime.now(timezone.utc).isoformat()
    output_path.write_text(json.dumps(fingerprints, indent=2, ensure_ascii=False), encoding="utf-8")

    logger.info(f"Fingerprints saved to {output_path}")
    return str(output_path)


def load_fingerprints(input_path: str = "") -> dict:
    """
    Load previously saved fingerprints from JSON.

    Args:
        input_path: Where to load from (default: outputs/_quinn_fingerprints.json)

    Returns:
        Dict of {"drl_template": fp, "signal_catalog": fp, "saved_at": ISO-8601}
        Returns empty dict if file doesn't exist.
    """
    if not input_path:
        input_path = Path(__file__).parent.parent / "outputs" / "_quinn_fingerprints.json"

    input_path = Path(input_path)
    if not input_path.exists():
        return {}

    try:
        data = json.loads(input_path.read_text(encoding="utf-8"))
        logger.info(f"Loaded fingerprints from {input_path}")
        return data
    except json.JSONDecodeError as e:
        logger.warning(f"Failed to load fingerprints from {input_path}: {e}")
        return {}
